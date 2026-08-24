from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from time import perf_counter

from openpyxl import load_workbook

from .common import (
    add_date_numeric_parse_warnings,
    build_headers,
    finish_result,
    is_empty_row,
    mark_truncated,
    mapped_values,
    make_issue,
    normalize_cell,
    set_header_summary,
    source_name_from,
    start_result,
)
from .errors import (
    CANONICAL_CELL_FORMULA_NO_CACHE,
    CANONICAL_HEADER_MISSING,
    CANONICAL_ROW_EMPTY,
    CANONICAL_SHEET_EMPTY,
    CANONICAL_SHEET_MISSING,
    CANONICAL_WORKBOOK_CORRUPTED,
    CANONICAL_WORKBOOK_EMPTY,
    CANONICAL_WORKBOOK_MISSING,
    CELL_KIND_BLANK,
    CELL_KIND_BOOLEAN,
    CELL_KIND_DATE,
    CELL_KIND_ERROR,
    CELL_KIND_NUMERIC,
    CELL_KIND_TEXT,
    EMPTY_ROW_SKIPPED,
    ERROR,
    FILE_EMPTY,
    FILE_NOT_FOUND,
    FILE_OPEN_FAILED,
    FATAL,
    ENTITY_CELL,
    ENTITY_HEADER,
    ENTITY_ROW,
    ENTITY_SHEET,
    ENTITY_WORKBOOK,
    INFO,
    ISSUE_KIND_STRUCTURAL,
    ROW_READ_FAILED,
    SHEET_NOT_FOUND,
    SCOPE_INPUT,
    SCOPE_ROW,
    SCOPE_TABLE,
)
from .models import Cell, Row, TabularResult


def read_xlsx_tabular(
    source: str | bytes | Path,
    *,
    source_name: str | None,
    header_row: int,
    sheet_name: str | None,
    sheet_index: int,
    header_mapping: dict[str, str] | None,
    max_rows: int | None,
    skip_empty_rows: bool,
    started_at: float | None = None,
) -> TabularResult:
    started_at = perf_counter() if started_at is None else started_at
    result = start_result(source_name_from(source, source_name), "xlsx")
    result.metadata.header_row_index = header_row

    raw = _read_bytes(source, result)
    if raw is None:
        return finish_result(result, started_at)
    if raw == b"":
        result.errors.append(
            make_issue(
                FILE_EMPTY,
                FATAL,
                SCOPE_INPUT,
                "XLSX file is empty.",
                entity=ENTITY_WORKBOOK,
                issue_kind=ISSUE_KIND_STRUCTURAL,
                canonical_code=CANONICAL_WORKBOOK_EMPTY,
            )
        )
        return finish_result(result, started_at)

    try:
        workbook = load_workbook(BytesIO(raw), data_only=True)
        workbook_formula = load_workbook(BytesIO(raw), data_only=False)
    except Exception as exc:
        result.errors.append(
            make_issue(
                FILE_OPEN_FAILED,
                FATAL,
                SCOPE_INPUT,
                "XLSX file could not be opened.",
                original_value=str(exc),
                entity=ENTITY_WORKBOOK,
                issue_kind=ISSUE_KIND_STRUCTURAL,
                canonical_code=CANONICAL_WORKBOOK_CORRUPTED,
            )
        )
        return finish_result(result, started_at)

    try:
        worksheet = _select_sheet(workbook, sheet_name, sheet_index, result)
        if worksheet is None:
            return finish_result(result, started_at)
        worksheet_formula = _mirror_sheet(workbook_formula, worksheet.title, sheet_index)
        result.metadata.selected_table = worksheet.title
        result.metadata.selected_sheet = worksheet.title

        values = list(worksheet.iter_rows())
        formula_rows = list(worksheet_formula.iter_rows())
        result.summary.total_physical_rows = len(values)
        if not values:
            result.errors.append(
                make_issue(
                    FILE_EMPTY,
                    FATAL,
                    SCOPE_TABLE,
                    "XLSX worksheet has no rows.",
                    entity=ENTITY_SHEET,
                    issue_kind=ISSUE_KIND_STRUCTURAL,
                    canonical_code=CANONICAL_SHEET_EMPTY,
                )
            )
            return finish_result(result, started_at)
        if header_row < 1 or header_row > len(values):
            result.errors.append(
                make_issue(
                    FILE_EMPTY,
                    FATAL,
                    SCOPE_INPUT,
                    "Header row is outside the worksheet row range.",
                    row=header_row,
                    suggestion="Use a header_row value that exists in the worksheet.",
                    entity=ENTITY_HEADER,
                    issue_kind=ISSUE_KIND_STRUCTURAL,
                    canonical_code=CANONICAL_HEADER_MISSING,
                )
            )
            return finish_result(result, started_at)

        header_values = [_cell_value(cell) for cell in values[header_row - 1]]
        headers, warnings = build_headers(header_values, header_mapping)
        result.headers = headers
        result.warnings.extend(warnings)
        set_header_summary(result)

        max_columns = max((max(len(row), len(formula_rows[index])) for index, row in enumerate(values)), default=0)

        for row_offset, value_row in enumerate(values[header_row:], start=header_row):
            source_row_index = row_offset + 1
            formula_row = formula_rows[row_offset] if row_offset < len(formula_rows) else ()
            try:
                cells, row_issues = _row_cells(value_row, formula_row, source_row_index, max_columns)
            except Exception as exc:
                result.errors.append(
                    make_issue(
                        ROW_READ_FAILED,
                        ERROR,
                        SCOPE_ROW,
                        "XLSX row could not be read.",
                        row=source_row_index,
                        original_value=str(exc),
                        entity=ENTITY_ROW,
                        issue_kind=ISSUE_KIND_STRUCTURAL,
                    )
                )
                continue
            row_values = [cell.value for cell in cells]
            result.warnings.extend(row_issues)
            if skip_empty_rows and is_empty_row(row_values):
                result.summary.skipped_rows += 1
                result.warnings.append(
                    make_issue(
                        EMPTY_ROW_SKIPPED,
                        INFO,
                        SCOPE_ROW,
                        "Empty row skipped.",
                        row=source_row_index,
                        entity=ENTITY_ROW,
                        issue_kind=ISSUE_KIND_STRUCTURAL,
                        canonical_code=CANONICAL_ROW_EMPTY,
                    )
                )
                continue
            if max_rows is not None and len(result.rows) >= max_rows:
                mark_truncated(result, source_row_index)
                break
            result.rows.append(
                Row(
                    row_index=len(result.rows) + 1,
                    source_row_index=source_row_index,
                    raw_values=row_values,
                    cells=cells,
                    mapped_values=mapped_values(headers, row_values),
                )
            )

        result.summary.total_data_rows = len(result.rows) + result.summary.skipped_rows
        add_date_numeric_parse_warnings(result)
        return finish_result(result, started_at)
    finally:
        workbook.close()
        workbook_formula.close()


def _read_bytes(source: str | bytes | Path, result: TabularResult) -> bytes | None:
    if isinstance(source, bytes):
        return source
    try:
        path = Path(source)
        if not path.exists():
            result.errors.append(
                make_issue(
                    FILE_NOT_FOUND,
                    FATAL,
                    SCOPE_INPUT,
                    "XLSX file was not found.",
                    entity=ENTITY_WORKBOOK,
                    issue_kind=ISSUE_KIND_STRUCTURAL,
                    canonical_code=CANONICAL_WORKBOOK_MISSING,
                )
            )
            return None
        return path.read_bytes()
    except OSError as exc:
        result.errors.append(
            make_issue(
                FILE_OPEN_FAILED,
                FATAL,
                SCOPE_INPUT,
                "XLSX file could not be opened.",
                original_value=str(exc),
                entity=ENTITY_WORKBOOK,
                issue_kind=ISSUE_KIND_STRUCTURAL,
            )
        )
        return None


def _select_sheet(workbook, sheet_name: str | None, sheet_index: int, result: TabularResult):
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            result.errors.append(
                make_issue(
                    SHEET_NOT_FOUND,
                    FATAL,
                    SCOPE_TABLE,
                    "Requested worksheet was not found.",
                    original_value=sheet_name,
                    suggestion="Use an existing sheet_name or omit it to use the first worksheet.",
                    entity=ENTITY_SHEET,
                    issue_kind=ISSUE_KIND_STRUCTURAL,
                    canonical_code=CANONICAL_SHEET_MISSING,
                )
            )
            return None
        return workbook[sheet_name]
    if sheet_index < 0 or sheet_index >= len(workbook.worksheets):
        result.errors.append(
            make_issue(
                SHEET_NOT_FOUND,
                FATAL,
                SCOPE_TABLE,
                "Requested worksheet index was not found.",
                original_value=sheet_index,
                suggestion="Use a zero-based sheet_index within the workbook sheet range.",
                entity=ENTITY_SHEET,
                issue_kind=ISSUE_KIND_STRUCTURAL,
                canonical_code=CANONICAL_SHEET_MISSING,
            )
        )
        return None
    return workbook.worksheets[sheet_index]


def _mirror_sheet(workbook, sheet_title: str, sheet_index: int):
    if sheet_title in workbook.sheetnames:
        return workbook[sheet_title]
    return workbook.worksheets[sheet_index]


def _row_cells(value_row, formula_row, source_row_index: int, max_columns: int) -> tuple[list[Cell], list]:
    cells: list[Cell] = []
    issues = []
    for column_index in range(1, max_columns + 1):
        value_cell = value_row[column_index - 1] if column_index - 1 < len(value_row) else None
        formula_cell = formula_row[column_index - 1] if column_index - 1 < len(formula_row) else None
        cell, issue = _build_cell(value_cell, formula_cell, source_row_index, column_index)
        cells.append(cell)
        if issue is not None:
            issues.append(issue)
    return cells, issues


def _build_cell(value_cell, formula_cell, source_row_index: int, column_index: int) -> tuple[Cell, object | None]:
    has_formula = bool(formula_cell is not None and getattr(formula_cell, "data_type", None) == "f")
    has_cached_value = None if not has_formula else value_cell is not None and _cell_value(value_cell) is not None
    value = _cell_value(value_cell)
    if has_formula and not has_cached_value:
        return (
            Cell(
                column_index=column_index,
                value=None,
                kind=CELL_KIND_BLANK,
                source_kind="formula",
                has_formula=True,
                has_cached_value=False,
            ),
            make_issue(
                code="FORMULA_NO_CACHED_VALUE",
                severity="warning",
                scope=SCOPE_ROW,
                message="Formula cell has no cached value.",
                row=source_row_index,
                column=column_index,
                original_value=getattr(formula_cell, "value", None),
                suggestion="Recalculate and save the workbook before import.",
                entity=ENTITY_CELL,
                issue_kind=ISSUE_KIND_STRUCTURAL,
                canonical_code=CANONICAL_CELL_FORMULA_NO_CACHE,
            ),
        )

    kind = _cell_kind(value_cell, formula_cell, value)
    source_kind = CELL_KIND_BLANK if value is None and not has_formula else ("formula" if has_formula else _source_kind(formula_cell, value_cell))
    return (
        Cell(
            column_index=column_index,
            value=value,
            kind=kind,
            source_kind=source_kind,
            has_formula=has_formula,
            has_cached_value=has_cached_value,
        ),
        None,
    )


def _cell_value(cell):
    if cell is None:
        return None
    value = cell.value
    if value is None and getattr(cell, "data_type", None) in {"inlineStr", "s", "str"}:
        return ""
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    return normalize_cell(value)


def _cell_kind(value_cell, formula_cell, value) -> str:
    source = formula_cell if formula_cell is not None else value_cell
    if value is None:
        return CELL_KIND_BLANK
    if source is not None and getattr(source, "data_type", None) == "e":
        return CELL_KIND_ERROR
    if isinstance(value, bool):
        return CELL_KIND_BOOLEAN
    if _is_date_cell(value_cell, formula_cell, value):
        return CELL_KIND_DATE
    if isinstance(value, (int, float)):
        return CELL_KIND_NUMERIC
    return CELL_KIND_TEXT


def _is_date_cell(value_cell, formula_cell, value) -> bool:
    if not isinstance(value, datetime):
        return False
    return bool(
        (value_cell is not None and getattr(value_cell, "is_date", False))
        or (formula_cell is not None and getattr(formula_cell, "is_date", False))
    )


def _source_kind(primary_cell, fallback_cell) -> str:
    source = primary_cell if primary_cell is not None else fallback_cell
    if source is None:
        return CELL_KIND_BLANK
    data_type = getattr(source, "data_type", None)
    if data_type in {"n", "d"}:
        return CELL_KIND_NUMERIC
    if data_type == "b":
        return CELL_KIND_BOOLEAN
    if data_type == "e":
        return CELL_KIND_ERROR
    return CELL_KIND_TEXT
